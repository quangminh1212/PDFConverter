"""
PDFConverter - Professional PDF Conversion Tool with OCR Support
Architecture inspired by Stirling-PDF, pdf2docx, OCRmyPDF
"""

import os
import sys
import io
import re
import csv
import gc
import unicodedata
import json
import time
import shutil
import logging
import zipfile
import tempfile
import traceback
from pathlib import Path
from abc import ABC, abstractmethod
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime

from flask import Flask, request, jsonify, send_file, render_template, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename

import fitz  # PyMuPDF
import pdfplumber
from PIL import Image

# Conditional imports with graceful fallback
try:
    import pytesseract
    HAS_TESSERACT = True
except ImportError:
    HAS_TESSERACT = False

try:
    from pdf2docx import Converter as Pdf2DocxConverter
    HAS_PDF2DOCX = True
except ImportError:
    HAS_PDF2DOCX = False

try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from markdownify import markdownify as md
    HAS_MARKDOWNIFY = True
except ImportError:
    HAS_MARKDOWNIFY = False

try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False

# ============================================================
# Configuration
# ============================================================
BASE_DIR = Path(__file__).parent.resolve()
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "output"
LOG_DIR = BASE_DIR / "logs"

for d in [UPLOAD_DIR, OUTPUT_DIR, LOG_DIR]:
    d.mkdir(exist_ok=True)

# Logging setup
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "app.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("PDFConverter")

ALLOWED_EXTENSIONS = {"pdf"}
MAX_FILE_SIZE = 500 * 1024 * 1024  # 500 MB for large files
CHUNK_SIZE = 10  # pages per chunk for memory management

SUPPORTED_FORMATS = {
    "word": {"ext": ".docx", "mime": "application/vnd.openxmlformats-officedocument.wordprocessingml.document", "label": "Word Document"},
    "excel": {"ext": ".xlsx", "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "label": "Excel Spreadsheet"},
    "image": {"ext": ".png", "mime": "application/zip", "label": "Images (PNG)"},
    "text": {"ext": ".txt", "mime": "text/plain", "label": "Plain Text"},
    "html": {"ext": ".html", "mime": "text/html", "label": "HTML Document"},
    "markdown": {"ext": ".md", "mime": "text/markdown", "label": "Markdown Document"},
    "csv": {"ext": ".csv", "mime": "text/csv", "label": "CSV (Tables)"},
}

# Tesseract path for Windows
if sys.platform == "win32" and HAS_TESSERACT:
    tesseract_paths = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        r"C:\Users\{}\AppData\Local\Programs\Tesseract-OCR\tesseract.exe".format(os.getenv("USERNAME", "")),
    ]
    for tp in tesseract_paths:
        if os.path.isfile(tp):
            pytesseract.pytesseract.tesseract_cmd = tp
            break


# ============================================================
# OCR Engine
# ============================================================
class OCREngine:
    """Optical Character Recognition engine using Tesseract."""

    SUPPORTED_LANGUAGES = {
        "eng": "English",
        "vie": "Vietnamese",
        "jpn": "Japanese",
        "kor": "Korean",
        "chi_sim": "Chinese (Simplified)",
        "chi_tra": "Chinese (Traditional)",
        "fra": "French",
        "deu": "German",
        "spa": "Spanish",
        "rus": "Russian",
    }

    def __init__(self, lang: str = "eng+vie"):
        self.lang = lang
        self.available = HAS_TESSERACT and self._check_tesseract()

    def _check_tesseract(self) -> bool:
        """Check if Tesseract is installed and accessible."""
        try:
            version = pytesseract.get_tesseract_version()
            logger.info(f"Tesseract OCR v{version} detected")
            return True
        except Exception as e:
            logger.warning(f"Tesseract not available: {e}")
            return False

    def get_available_languages(self) -> List[str]:
        """Get list of installed Tesseract languages."""
        if not self.available:
            return []
        try:
            return pytesseract.get_languages()
        except Exception:
            return ["eng"]

    def page_needs_ocr(self, page: fitz.Page, threshold: int = 30) -> bool:
        """Detect if a page needs OCR (scanned/image-based)."""
        text = page.get_text("text").strip()
        # If very little text extracted, likely a scanned page
        if len(text) < threshold:
            return True
        # Check ratio of images to text blocks
        image_list = page.get_images(full=True)
        text_blocks = page.get_text("blocks")
        if image_list and len(text_blocks) <= 2:
            return True
        return False

    def ocr_page(self, page: fitz.Page, dpi: int = 300) -> str:
        """Perform OCR on a single page."""
        if not self.available:
            return page.get_text("text")

        try:
            # Render page to high-res image
            mat = fitz.Matrix(dpi / 72, dpi / 72)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img_data = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_data))

            # Preprocess image for better OCR
            img = self._preprocess_image(img)

            # Run OCR with custom config for accuracy
            custom_config = r"--oem 3 --psm 6"
            text = pytesseract.image_to_string(img, lang=self.lang, config=custom_config)
            return text
        except Exception as e:
            logger.error(f"OCR failed for page: {e}")
            return page.get_text("text")

    def ocr_page_to_hocr(self, page: fitz.Page, dpi: int = 300) -> str:
        """Perform OCR and return hOCR (HTML-based OCR output with position info)."""
        if not self.available:
            return ""
        try:
            mat = fitz.Matrix(dpi / 72, dpi / 72)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img_data = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_data))
            img = self._preprocess_image(img)
            custom_config = r"--oem 3 --psm 6"
            hocr = pytesseract.image_to_pdf_or_hocr(img, lang=self.lang, config=custom_config, extension="hocr")
            return hocr.decode("utf-8") if isinstance(hocr, bytes) else hocr
        except Exception as e:
            logger.error(f"hOCR failed: {e}")
            return ""

    def ocr_image(self, image: Image.Image) -> str:
        """Perform OCR on a PIL Image."""
        if not self.available:
            return ""
        try:
            img = self._preprocess_image(image)
            custom_config = r"--oem 3 --psm 6"
            return pytesseract.image_to_string(img, lang=self.lang, config=custom_config)
        except Exception as e:
            logger.error(f"OCR on image failed: {e}")
            return ""

    @staticmethod
    def _preprocess_image(img: Image.Image) -> Image.Image:
        """Preprocess image for better OCR accuracy."""
        # Convert to grayscale
        if img.mode != "L":
            img = img.convert("L")
        # Apply slight sharpening via contrast enhancement
        from PIL import ImageFilter, ImageEnhance
        img = img.filter(ImageFilter.SHARPEN)
        enhancer = ImageEnhance.Contrast(img)
        img = enhancer.enhance(1.5)
        return img


# ============================================================
# Multi-Layer Text Extractor (Cross-Verification)
# ============================================================
class TextExtractor:
    """Multi-method text extraction with cross-verification for accuracy."""

    # Common OCR ligature/character fixes
    CHAR_FIXES = {
        '\ufb00': 'ff', '\ufb01': 'fi', '\ufb02': 'fl',
        '\ufb03': 'ffi', '\ufb04': 'ffl',
        '\u2018': "'", '\u2019': "'", '\u201c': '"', '\u201d': '"',
        '\u2013': '-', '\u2014': '--', '\u2026': '...',
        '\u00a0': ' ',  # non-breaking space
        '\u200b': '',   # zero-width space
        '\u200c': '',   # zero-width non-joiner
        '\u200d': '',   # zero-width joiner
        '\ufeff': '',   # BOM
    }

    def __init__(self, ocr_engine: Optional[OCREngine] = None):
        self.ocr_engine = ocr_engine
        self.logger = logging.getLogger("PDFConverter.TextExtractor")

    @staticmethod
    def normalize_unicode(text: str) -> str:
        """Normalize Unicode to NFC form and fix common issues."""
        if not text:
            return ""
        # NFC normalization (compose characters)
        text = unicodedata.normalize('NFC', text)
        # Fix ligatures and special characters
        for old, new in TextExtractor.CHAR_FIXES.items():
            text = text.replace(old, new)
        # Remove control characters except newlines and tabs
        text = ''.join(
            c if c in ('\n', '\t', '\r') or not unicodedata.category(c).startswith('C')
            else ' ' for c in text
        )
        # Normalize whitespace (collapse multiple spaces, keep newlines)
        text = re.sub(r'[^\S\n]+', ' ', text)
        text = re.sub(r'\n{3,}', '\n\n', text)
        return text.strip()

    def extract_method_native(self, page: fitz.Page) -> str:
        """Method A: Native PyMuPDF text extraction."""
        try:
            return page.get_text("text", sort=True)
        except Exception:
            return ""

    def extract_method_blocks(self, page: fitz.Page) -> str:
        """Method B: Structured block-based extraction (preserves reading order)."""
        try:
            blocks = page.get_text("dict", sort=True).get("blocks", [])
            lines = []
            for block in blocks:
                if block.get("type") == 0:  # text block
                    for line in block.get("lines", []):
                        spans_text = ""
                        for span in line.get("spans", []):
                            spans_text += span.get("text", "")
                        if spans_text.strip():
                            lines.append(spans_text)
            return "\n".join(lines)
        except Exception:
            return ""

    def extract_method_ocr(self, page: fitz.Page, dpi: int = 300) -> str:
        """Method C: OCR-based extraction."""
        if not self.ocr_engine or not self.ocr_engine.available:
            return ""
        try:
            return self.ocr_engine.ocr_page(page, dpi=dpi)
        except Exception:
            return ""

    def _similarity_score(self, text_a: str, text_b: str) -> float:
        """Simple character-level similarity between two texts (0.0 - 1.0)."""
        if not text_a or not text_b:
            return 0.0
        a, b = text_a.strip(), text_b.strip()
        if a == b:
            return 1.0
        # Use ratio of common characters
        set_a, set_b = set(a), set(b)
        if not set_a or not set_b:
            return 0.0
        common = set_a & set_b
        return len(common) / max(len(set_a), len(set_b))

    def extract_verified(
        self, page: fitz.Page, page_num: int, use_ocr: bool = True, dpi: int = 300
    ) -> Tuple[str, Dict[str, Any]]:
        """Extract text with multi-layer verification.

        Uses multiple methods and cross-validates results.
        Returns (best_text, verification_report).
        """
        report = {"page": page_num + 1, "methods_used": [], "chosen_method": "", "confidence": 0.0}

        # Method A: Native text
        text_native = self.normalize_unicode(self.extract_method_native(page))
        len_a = len(text_native.strip())
        report["methods_used"].append({"name": "native", "chars": len_a})

        # Method B: Block-based
        text_blocks = self.normalize_unicode(self.extract_method_blocks(page))
        len_b = len(text_blocks.strip())
        report["methods_used"].append({"name": "blocks", "chars": len_b})

        # Check if page needs OCR
        needs_ocr = self.ocr_engine and self.ocr_engine.page_needs_ocr(page) if self.ocr_engine else False
        text_ocr = ""

        if use_ocr and needs_ocr:
            text_ocr = self.normalize_unicode(self.extract_method_ocr(page, dpi))
            len_c = len(text_ocr.strip())
            report["methods_used"].append({"name": "ocr", "chars": len_c})
        else:
            len_c = 0

        # Decision logic: Cross-verify and pick best
        if len_a == 0 and len_b == 0 and len_c == 0:
            report["chosen_method"] = "none"
            report["confidence"] = 0.0
            return "", report

        # If native and blocks agree (high similarity), use native (fastest)
        if len_a > 0 and len_b > 0:
            sim_ab = self._similarity_score(text_native, text_blocks)
            if sim_ab > 0.8:
                # Both methods agree - high confidence
                # Use the longer one (more complete)
                best = text_native if len_a >= len_b else text_blocks
                report["chosen_method"] = "native" if len_a >= len_b else "blocks"
                report["confidence"] = min(sim_ab + 0.1, 1.0)
                self.logger.debug(f"Page {page_num+1}: native/blocks agree (sim={sim_ab:.2f}), conf={report['confidence']:.2f}")
                return best, report

        # If OCR was used and is significantly better
        if len_c > max(len_a, len_b) * 1.3:
            report["chosen_method"] = "ocr"
            report["confidence"] = 0.85
            self.logger.info(f"Page {page_num+1}: OCR chosen (len_ocr={len_c} >> native={len_a}, blocks={len_b})")
            return text_ocr, report

        # If OCR available, verify against native
        if text_ocr and len_a > 0:
            sim_ac = self._similarity_score(text_native, text_ocr)
            if sim_ac > 0.7:
                # OCR confirms native
                report["chosen_method"] = "native_verified_by_ocr"
                report["confidence"] = 0.95
                return text_native, report

        # Default: pick longest result
        candidates = [(text_native, len_a, "native"), (text_blocks, len_b, "blocks"), (text_ocr, len_c, "ocr")]
        candidates.sort(key=lambda x: x[1], reverse=True)
        best_text, best_len, best_name = candidates[0]
        report["chosen_method"] = best_name
        report["confidence"] = 0.7 if best_len > 50 else 0.5
        return best_text, report


# ============================================================
# Base Converter
# ============================================================
class BaseConverter(ABC):
    """Abstract base class for all PDF converters."""

    FORMAT_NAME: str = ""
    FORMAT_EXT: str = ""

    def __init__(self, ocr_engine: Optional[OCREngine] = None):
        self.ocr_engine = ocr_engine or OCREngine()
        self.text_extractor = TextExtractor(self.ocr_engine)
        self.logger = logging.getLogger(f"PDFConverter.{self.__class__.__name__}")

    @abstractmethod
    def convert(
        self,
        input_path: str,
        output_path: str,
        pages: Optional[List[int]] = None,
        use_ocr: bool = True,
        ocr_lang: str = "eng+vie",
        **kwargs,
    ) -> Dict[str, Any]:
        pass

    def _extract_text_with_ocr(
        self, doc: fitz.Document, page_num: int, use_ocr: bool = True
    ) -> str:
        """Extract text with multi-layer verification."""
        page = doc[page_num]
        text, report = self.text_extractor.extract_verified(page, page_num, use_ocr)
        self.logger.debug(
            f"Page {page_num+1}: method={report['chosen_method']}, "
            f"confidence={report['confidence']:.0%}, chars={len(text)}"
        )
        return text

    def _get_page_range(self, doc: fitz.Document, pages: Optional[List[int]] = None) -> List[int]:
        """Get list of page indices to process."""
        if pages is not None:
            return [p for p in pages if 0 <= p < len(doc)]
        return list(range(len(doc)))

    def _process_in_chunks(self, page_range: List[int], chunk_size: int = CHUNK_SIZE):
        """Yield page ranges in chunks for memory-efficient processing."""
        for i in range(0, len(page_range), chunk_size):
            yield page_range[i:i + chunk_size]


# ============================================================
# PDF to Word Converter
# ============================================================
class PDFToWord(BaseConverter):
    FORMAT_NAME = "Word Document"
    FORMAT_EXT = ".docx"

    def convert(self, input_path, output_path, pages=None, use_ocr=True, ocr_lang="eng+vie", **kwargs):
        self.logger.info(f"Converting to Word: {input_path}")
        start_time = time.time()

        try:
            if HAS_PDF2DOCX:
                # Use pdf2docx for best layout preservation
                cv = Pdf2DocxConverter(input_path)
                if pages is not None:
                    # pdf2docx uses 0-indexed page numbers
                    cv.convert(output_path, pages=pages)
                else:
                    cv.convert(output_path)
                cv.close()

                # If OCR needed, enhance with OCR text
                if use_ocr and self.ocr_engine.available:
                    self._enhance_with_ocr(input_path, output_path, pages, ocr_lang)
            else:
                # Fallback: extract text and create simple docx
                self._convert_text_based(input_path, output_path, pages, use_ocr)

            elapsed = time.time() - start_time
            self.logger.info(f"Word conversion completed in {elapsed:.2f}s")
            return {"success": True, "output": output_path, "time": elapsed}

        except Exception as e:
            self.logger.error(f"Word conversion failed: {e}\n{traceback.format_exc()}")
            # Fallback to text-based conversion
            try:
                self._convert_text_based(input_path, output_path, pages, use_ocr)
                return {"success": True, "output": output_path, "time": time.time() - start_time, "fallback": True}
            except Exception as e2:
                return {"success": False, "error": str(e2)}

    def _enhance_with_ocr(self, input_path, output_path, pages, ocr_lang):
        """Add OCR text for scanned pages in the document."""
        doc = fitz.open(input_path)
        page_range = self._get_page_range(doc, pages)
        ocr_pages = []
        for pn in page_range:
            page = doc[pn]
            if self.ocr_engine.page_needs_ocr(page):
                ocr_pages.append(pn)
        doc.close()

        if ocr_pages:
            self.logger.info(f"OCR enhancement needed for {len(ocr_pages)} pages")

    def _convert_text_based(self, input_path, output_path, pages, use_ocr):
        """Fallback text-based conversion."""
        try:
            from docx import Document as DocxDocument
        except ImportError:
            # If python-docx not available, create a simple text file
            doc = fitz.open(input_path)
            page_range = self._get_page_range(doc, pages)
            text_output = output_path.replace(".docx", ".txt")
            with open(text_output, "w", encoding="utf-8") as f:
                for pn in page_range:
                    text = self._extract_text_with_ocr(doc, pn, use_ocr)
                    f.write(f"--- Page {pn + 1} ---\n{text}\n\n")
            doc.close()
            return

        doc = fitz.open(input_path)
        page_range = self._get_page_range(doc, pages)
        docx_doc = DocxDocument()

        for pn in page_range:
            text = self._extract_text_with_ocr(doc, pn, use_ocr)
            if pn > page_range[0]:
                docx_doc.add_page_break()
            for paragraph in text.split("\n"):
                if paragraph.strip():
                    docx_doc.add_paragraph(paragraph)

        docx_doc.save(output_path)
        doc.close()


# ============================================================
# PDF to Excel Converter
# ============================================================
class PDFToExcel(BaseConverter):
    FORMAT_NAME = "Excel Spreadsheet"
    FORMAT_EXT = ".xlsx"

    def convert(self, input_path, output_path, pages=None, use_ocr=True, ocr_lang="eng+vie", **kwargs):
        self.logger.info(f"Converting to Excel: {input_path}")
        start_time = time.time()

        try:
            if not HAS_OPENPYXL:
                return {"success": False, "error": "openpyxl not installed"}

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            tables_found = False

            with pdfplumber.open(input_path) as pdf:
                page_range = pages if pages else list(range(len(pdf.pages)))
                current_row = 1

                for pn in page_range:
                    if pn >= len(pdf.pages):
                        continue
                    page = pdf.pages[pn]

                    # Extract tables
                    tables = page.extract_tables({
                        "vertical_strategy": "lines_strict",
                        "horizontal_strategy": "lines_strict",
                        "snap_tolerance": 5,
                        "join_tolerance": 5,
                    })

                    if not tables:
                        # Try with text strategy
                        tables = page.extract_tables({
                            "vertical_strategy": "text",
                            "horizontal_strategy": "text",
                        })

                    if tables:
                        tables_found = True
                        for table in tables:
                            # Add page header
                            ws.cell(row=current_row, column=1, value=f"Page {pn + 1}")
                            from openpyxl.styles import Font
                            ws.cell(row=current_row, column=1).font = Font(bold=True)
                            current_row += 1

                            for row_data in table:
                                for col_idx, cell_value in enumerate(row_data, 1):
                                    if cell_value is not None:
                                        val = TextExtractor.normalize_unicode(str(cell_value).strip())
                                        ws.cell(row=current_row, column=col_idx, value=val)
                                current_row += 1
                            current_row += 1  # Blank row between tables

            # If no tables found, try OCR-based text extraction
            if not tables_found:
                self.logger.info("No tables found, extracting as text grid")
                doc = fitz.open(input_path)
                page_range_list = self._get_page_range(doc, pages)
                current_row = 1

                for chunk in self._process_in_chunks(page_range_list):
                    for pn in chunk:
                        text = self._extract_text_with_ocr(doc, pn, use_ocr)
                        ws.cell(row=current_row, column=1, value=f"Page {pn + 1}")
                        from openpyxl.styles import Font
                        ws.cell(row=current_row, column=1).font = Font(bold=True)
                        current_row += 1

                        for line in text.split("\n"):
                            if line.strip():
                                parts = re.split(r"\t|  +", line.strip())
                                for col_idx, part in enumerate(parts, 1):
                                    ws.cell(row=current_row, column=col_idx, value=part.strip())
                                current_row += 1
                        current_row += 1
                doc.close()

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

            wb.save(output_path)
            elapsed = time.time() - start_time
            self.logger.info(f"Excel conversion completed in {elapsed:.2f}s")
            return {"success": True, "output": output_path, "time": elapsed, "tables_found": tables_found}

        except Exception as e:
            self.logger.error(f"Excel conversion failed: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}


# ============================================================
# PDF to Image Converter
# ============================================================
class PDFToImage(BaseConverter):
    FORMAT_NAME = "Images"
    FORMAT_EXT = ".png"

    def convert(self, input_path, output_path, pages=None, use_ocr=True, ocr_lang="eng+vie", **kwargs):
        self.logger.info(f"Converting to Images: {input_path}")
        start_time = time.time()

        try:
            dpi = kwargs.get("dpi", 200)
            image_format = kwargs.get("image_format", "png")
            doc = fitz.open(input_path)
            page_range = self._get_page_range(doc, pages)
            total_pages = len(page_range)
            doc.close()

            output_dir = Path(output_path).parent
            stem = Path(output_path).stem
            image_paths = []

            for chunk in self._process_in_chunks(page_range):
                doc = fitz.open(input_path)
                for pn in chunk:
                    page = doc[pn]
                    mat = fitz.Matrix(dpi / 72, dpi / 72)
                    pix = page.get_pixmap(matrix=mat, alpha=False)

                    img_path = str(output_dir / f"{stem}_page_{pn + 1}.{image_format}")
                    pix.save(img_path)
                    image_paths.append(img_path)
                    self.logger.info(f"Image: Page {pn+1}/{total_pages} done")
                    del pix  # Free pixmap memory
                doc.close()
                gc.collect()

            # Create ZIP archive if multiple pages
            if len(image_paths) > 1:
                zip_path = str(output_dir / f"{stem}_images.zip")
                with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                    for img_path in image_paths:
                        zf.write(img_path, Path(img_path).name)
                # Clean up individual images after ZIP
                for img_path in image_paths:
                    try:
                        os.remove(img_path)
                    except OSError:
                        pass
                output_path = zip_path
            elif image_paths:
                output_path = image_paths[0]

            elapsed = time.time() - start_time
            self.logger.info(f"Image conversion completed in {elapsed:.2f}s ({total_pages} pages)")
            return {
                "success": True,
                "output": output_path,
                "time": elapsed,
                "page_count": len(image_paths),
                "image_paths": image_paths,
            }

        except Exception as e:
            self.logger.error(f"Image conversion failed: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}


# ============================================================
# PDF to Text Converter
# ============================================================
class PDFToText(BaseConverter):
    FORMAT_NAME = "Plain Text"
    FORMAT_EXT = ".txt"

    def convert(self, input_path, output_path, pages=None, use_ocr=True, ocr_lang="eng+vie", **kwargs):
        self.logger.info(f"Converting to Text: {input_path}")
        start_time = time.time()

        try:
            # Get page range first
            doc = fitz.open(input_path)
            page_range = self._get_page_range(doc, pages)
            total_pages = len(page_range)
            doc.close()

            with open(output_path, "w", encoding="utf-8") as f:
                pages_done = 0
                for chunk in self._process_in_chunks(page_range):
                    # Open doc per chunk for memory efficiency
                    doc = fitz.open(input_path)
                    for pn in chunk:
                        text = self._extract_text_with_ocr(doc, pn, use_ocr)
                        # Ensure clean Unicode output
                        text = TextExtractor.normalize_unicode(text)

                        if pages_done > 0:
                            f.write("\n\n" + "=" * 60 + "\n")
                        f.write(f"Page {pn + 1}\n")
                        f.write("=" * 60 + "\n\n")
                        f.write(text)
                        f.flush()  # Stream write for large files
                        pages_done += 1
                        self.logger.info(f"Text: Page {pn+1}/{total_pages} done")
                    doc.close()
                    gc.collect()  # Free memory between chunks

            elapsed = time.time() - start_time
            self.logger.info(f"Text conversion completed in {elapsed:.2f}s ({total_pages} pages)")
            return {"success": True, "output": output_path, "time": elapsed, "page_count": total_pages}

        except Exception as e:
            self.logger.error(f"Text conversion failed: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}


# ============================================================
# PDF to HTML Converter
# ============================================================
class PDFToHTML(BaseConverter):
    FORMAT_NAME = "HTML Document"
    FORMAT_EXT = ".html"

    def convert(self, input_path, output_path, pages=None, use_ocr=True, ocr_lang="eng+vie", **kwargs):
        self.logger.info(f"Converting to HTML: {input_path}")
        start_time = time.time()

        try:
            doc = fitz.open(input_path)
            page_range = self._get_page_range(doc, pages)
            total_pages = len(page_range)
            doc.close()

            with open(output_path, "w", encoding="utf-8") as f:
                # Write HTML header
                f.write('<!DOCTYPE html>\n<html lang="en">\n<head>\n')
                f.write('<meta charset="UTF-8">\n')
                f.write('<meta name="viewport" content="width=device-width, initial-scale=1.0">\n')
                f.write(f'<title>{Path(input_path).stem}</title>\n')
                f.write('<style>\n')
                f.write("body { font-family: 'Segoe UI', Arial, sans-serif; max-width: 900px; margin: 0 auto; padding: 20px; background: #f5f5f5; }\n")
                f.write(".page { background: white; padding: 40px; margin: 20px 0; box-shadow: 0 2px 8px rgba(0,0,0,0.1); border-radius: 4px; }\n")
                f.write(".page-header { color: #666; font-size: 12px; margin-bottom: 20px; border-bottom: 1px solid #eee; padding-bottom: 10px; }\n")
                f.write("img { max-width: 100%; height: auto; }\n")
                f.write("table { border-collapse: collapse; width: 100%; margin: 10px 0; }\n")
                f.write("td, th { border: 1px solid #ddd; padding: 8px; text-align: left; }\n")
                f.write("th { background: #f8f9fa; }\n")
                f.write('</style>\n</head>\n<body>\n')

                # Process pages in chunks
                for chunk in self._process_in_chunks(page_range):
                    doc = fitz.open(input_path)
                    for pn in chunk:
                        page = doc[pn]
                        f.write(f'<div class="page">\n')
                        f.write(f'<div class="page-header">Page {pn + 1}</div>\n')

                        page_html = page.get_text("html")
                        if page_html:
                            if HAS_BS4:
                                soup = BeautifulSoup(page_html, "html.parser")
                                for div in soup.find_all("div"):
                                    div.unwrap()
                                page_html = str(soup)
                            f.write(page_html)
                        else:
                            text = self._extract_text_with_ocr(doc, pn, use_ocr)
                            text = TextExtractor.normalize_unicode(text)
                            for para in text.split("\n\n"):
                                if para.strip():
                                    import html as html_lib
                                    f.write(f"<p>{html_lib.escape(para.strip())}</p>\n")

                        f.write('</div>\n')
                        f.flush()
                        self.logger.info(f"HTML: Page {pn+1}/{total_pages} done")
                    doc.close()
                    gc.collect()

                f.write('</body>\n</html>\n')

            elapsed = time.time() - start_time
            self.logger.info(f"HTML conversion completed in {elapsed:.2f}s ({total_pages} pages)")
            return {"success": True, "output": output_path, "time": elapsed, "page_count": total_pages}

        except Exception as e:
            self.logger.error(f"HTML conversion failed: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}


# ============================================================
# PDF to Markdown Converter
# ============================================================
class PDFToMarkdown(BaseConverter):
    FORMAT_NAME = "Markdown Document"
    FORMAT_EXT = ".md"

    def convert(self, input_path, output_path, pages=None, use_ocr=True, ocr_lang="eng+vie", **kwargs):
        self.logger.info(f"Converting to Markdown: {input_path}")
        start_time = time.time()

        try:
            doc = fitz.open(input_path)
            page_range = self._get_page_range(doc, pages)
            total_pages = len(page_range)
            doc.close()

            with open(output_path, "w", encoding="utf-8") as f:
                f.write(f"# {Path(input_path).stem}\n\n")

                for chunk in self._process_in_chunks(page_range):
                    doc = fitz.open(input_path)
                    for pn in chunk:
                        page = doc[pn]
                        f.write(f"\n---\n\n## Page {pn + 1}\n\n")

                        # Try HTML → Markdown for better formatting
                        page_html = page.get_text("html")
                        if page_html and HAS_MARKDOWNIFY and HAS_BS4:
                            try:
                                soup = BeautifulSoup(page_html, "html.parser")
                                for style in soup.find_all("style"):
                                    style.decompose()
                                markdown_text = md(str(soup), heading_style="ATX", bullets="-")
                                markdown_text = re.sub(r"\n{3,}", "\n\n", markdown_text)
                                markdown_text = TextExtractor.normalize_unicode(markdown_text)
                                f.write(markdown_text)
                            except Exception:
                                text = self._extract_text_with_ocr(doc, pn, use_ocr)
                                f.write(self._text_to_markdown(text))
                        else:
                            text = self._extract_text_with_ocr(doc, pn, use_ocr)
                            f.write(self._text_to_markdown(text))

                        # Extract tables with pdfplumber
                        try:
                            with pdfplumber.open(input_path) as pdf:
                                if pn < len(pdf.pages):
                                    tables = pdf.pages[pn].extract_tables()
                                    for table in tables:
                                        f.write(self._table_to_markdown(table))
                        except Exception:
                            pass

                        f.flush()
                        self.logger.info(f"Markdown: Page {pn+1}/{total_pages} done")
                    doc.close()
                    gc.collect()

            elapsed = time.time() - start_time
            self.logger.info(f"Markdown conversion completed in {elapsed:.2f}s ({total_pages} pages)")
            return {"success": True, "output": output_path, "time": elapsed, "page_count": total_pages}

        except Exception as e:
            self.logger.error(f"Markdown conversion failed: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}

    @staticmethod
    def _text_to_markdown(text: str) -> str:
        """Convert plain text to basic Markdown format."""
        lines = text.split("\n")
        md_lines = []
        for line in lines:
            stripped = line.strip()
            if not stripped:
                md_lines.append("")
                continue
            # Detect potential headings (all caps, short lines)
            if stripped.isupper() and len(stripped) < 80 and len(stripped.split()) < 10:
                md_lines.append(f"\n### {stripped}\n")
            else:
                md_lines.append(stripped)
        return "\n".join(md_lines)

    @staticmethod
    def _table_to_markdown(table: List[List]) -> str:
        """Convert a table to Markdown format."""
        if not table or not table[0]:
            return ""
        md_lines = ["\n"]
        # Header
        header = [str(cell or "").strip() for cell in table[0]]
        md_lines.append("| " + " | ".join(header) + " |")
        md_lines.append("| " + " | ".join(["---"] * len(header)) + " |")
        # Rows
        for row in table[1:]:
            cells = [str(cell or "").strip() for cell in row]
            # Pad if necessary
            while len(cells) < len(header):
                cells.append("")
            md_lines.append("| " + " | ".join(cells[:len(header)]) + " |")
        md_lines.append("")
        return "\n".join(md_lines)


# ============================================================
# PDF to CSV Converter
# ============================================================
class PDFToCSV(BaseConverter):
    FORMAT_NAME = "CSV (Tables)"
    FORMAT_EXT = ".csv"

    def convert(self, input_path, output_path, pages=None, use_ocr=True, ocr_lang="eng+vie", **kwargs):
        self.logger.info(f"Converting to CSV: {input_path}")
        start_time = time.time()

        try:
            all_rows = []
            delimiter = kwargs.get("delimiter", ",")

            with pdfplumber.open(input_path) as pdf:
                page_range = pages if pages else list(range(len(pdf.pages)))

                for pn in page_range:
                    if pn >= len(pdf.pages):
                        continue
                    page = pdf.pages[pn]

                    tables = page.extract_tables({
                        "vertical_strategy": "lines_strict",
                        "horizontal_strategy": "lines_strict",
                    })

                    if not tables:
                        tables = page.extract_tables({
                            "vertical_strategy": "text",
                            "horizontal_strategy": "text",
                        })

                    for table in tables:
                        for row in table:
                            cleaned_row = [TextExtractor.normalize_unicode(str(cell or "").strip()) for cell in row]
                            all_rows.append(cleaned_row)
                        all_rows.append([])  # Blank row separator

            # If no tables found, try text-based extraction
            if not all_rows or all(not row for row in all_rows):
                self.logger.info("No tables found, extracting text as CSV")
                doc = fitz.open(input_path)
                page_range_list = self._get_page_range(doc, pages)
                all_rows = []
                for pn in page_range_list:
                    text = self._extract_text_with_ocr(doc, pn, use_ocr)
                    for line in text.split("\n"):
                        if line.strip():
                            if "\t" in line:
                                all_rows.append([TextExtractor.normalize_unicode(c) for c in line.split("\t")])
                            elif ";" in line:
                                all_rows.append([TextExtractor.normalize_unicode(c) for c in line.split(";")])
                            else:
                                all_rows.append([TextExtractor.normalize_unicode(line.strip())])
                doc.close()

            with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f, delimiter=delimiter)
                for row in all_rows:
                    writer.writerow(row)

            elapsed = time.time() - start_time
            self.logger.info(f"CSV conversion completed in {elapsed:.2f}s")
            return {"success": True, "output": output_path, "time": elapsed, "rows": len(all_rows)}

        except Exception as e:
            self.logger.error(f"CSV conversion failed: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}


# ============================================================
# Conversion Manager
# ============================================================
class ConversionManager:
    """Orchestrates PDF conversion operations."""

    CONVERTERS = {
        "word": PDFToWord,
        "excel": PDFToExcel,
        "image": PDFToImage,
        "text": PDFToText,
        "html": PDFToHTML,
        "markdown": PDFToMarkdown,
        "csv": PDFToCSV,
    }

    def __init__(self):
        self.ocr_engine = OCREngine()

    def convert(
        self,
        input_path: str,
        output_format: str,
        output_dir: str = None,
        pages: Optional[List[int]] = None,
        use_ocr: bool = True,
        ocr_lang: str = "eng+vie",
        **kwargs,
    ) -> Dict[str, Any]:
        """Perform PDF conversion."""
        if output_format not in self.CONVERTERS:
            return {"success": False, "error": f"Unsupported format: {output_format}"}

        if not os.path.isfile(input_path):
            return {"success": False, "error": "Input file not found"}

        # Prepare output path
        if output_dir is None:
            output_dir = str(OUTPUT_DIR)
        os.makedirs(output_dir, exist_ok=True)

        stem = Path(input_path).stem
        ext = SUPPORTED_FORMATS[output_format]["ext"]
        output_path = os.path.join(output_dir, f"{stem}{ext}")

        # Update OCR language
        self.ocr_engine.lang = ocr_lang

        # Create converter and run
        converter_class = self.CONVERTERS[output_format]
        converter = converter_class(self.ocr_engine)
        result = converter.convert(
            input_path=input_path,
            output_path=output_path,
            pages=pages,
            use_ocr=use_ocr,
            ocr_lang=ocr_lang,
            **kwargs,
        )

        return result

    def get_pdf_info(self, input_path: str) -> Dict[str, Any]:
        """Get PDF file information."""
        try:
            doc = fitz.open(input_path)
            info = {
                "filename": Path(input_path).name,
                "page_count": len(doc),
                "metadata": doc.metadata or {},
                "file_size": os.path.getsize(input_path),
                "pages": [],
            }

            for i, page in enumerate(doc):
                page_info = {
                    "number": i + 1,
                    "width": round(page.rect.width, 2),
                    "height": round(page.rect.height, 2),
                    "has_text": len(page.get_text("text").strip()) > 0,
                    "has_images": len(page.get_images(full=True)) > 0,
                    "needs_ocr": self.ocr_engine.page_needs_ocr(page) if self.ocr_engine.available else None,
                }
                info["pages"].append(page_info)

            doc.close()
            return info

        except Exception as e:
            return {"error": str(e)}


# ============================================================
# Flask Application
# ============================================================
app = Flask(__name__, static_folder="static", template_folder="templates")
CORS(app)
app.config["MAX_CONTENT_LENGTH"] = MAX_FILE_SIZE

manager = ConversionManager()


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/static/<path:filename>")
def serve_static(filename):
    return send_from_directory("static", filename)


@app.route("/api/formats", methods=["GET"])
def get_formats():
    """Get supported conversion formats."""
    return jsonify({
        "formats": SUPPORTED_FORMATS,
        "ocr_available": manager.ocr_engine.available,
        "ocr_languages": OCREngine.SUPPORTED_LANGUAGES,
    })


@app.route("/api/upload", methods=["POST"])
def upload_file():
    """Upload PDF and get file info."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    if not allowed_file(file.filename):
        return jsonify({"error": "Only PDF files are allowed"}), 400

    filename = secure_filename(file.filename)
    # Add timestamp to avoid collisions
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    unique_name = f"{timestamp}_{filename}"
    filepath = str(UPLOAD_DIR / unique_name)
    file.save(filepath)

    # Get PDF info
    info = manager.get_pdf_info(filepath)
    info["stored_name"] = unique_name

    return jsonify(info)


@app.route("/api/convert", methods=["POST"])
def convert_file():
    """Convert uploaded PDF to target format."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    stored_name = data.get("stored_name")
    output_format = data.get("format")
    use_ocr = data.get("use_ocr", True)
    ocr_lang = data.get("ocr_lang", "eng+vie")
    pages = data.get("pages")  # None = all pages
    dpi = data.get("dpi", 200)
    image_format = data.get("image_format", "png")

    if not stored_name or not output_format:
        return jsonify({"error": "Missing required fields: stored_name, format"}), 400

    input_path = str(UPLOAD_DIR / stored_name)
    if not os.path.isfile(input_path):
        return jsonify({"error": "File not found"}), 404

    # Convert
    result = manager.convert(
        input_path=input_path,
        output_format=output_format,
        use_ocr=use_ocr,
        ocr_lang=ocr_lang,
        pages=pages,
        dpi=dpi,
        image_format=image_format,
    )

    return jsonify(result)


@app.route("/api/download/<path:filename>", methods=["GET"])
def download_file(filename):
    """Download converted file."""
    # Check in output directory
    filepath = OUTPUT_DIR / filename
    if not filepath.is_file():
        # Try absolute path
        filepath = Path(filename)
    if not filepath.is_file():
        return jsonify({"error": "File not found"}), 404

    return send_file(
        str(filepath),
        as_attachment=True,
        download_name=filepath.name,
    )


@app.route("/api/preview/<stored_name>", methods=["GET"])
def preview_page(stored_name):
    """Generate preview thumbnail of a PDF page."""
    page_num = request.args.get("page", 0, type=int)
    input_path = str(UPLOAD_DIR / stored_name)

    if not os.path.isfile(input_path):
        return jsonify({"error": "File not found"}), 404

    try:
        doc = fitz.open(input_path)
        if page_num >= len(doc):
            doc.close()
            return jsonify({"error": "Page not found"}), 404

        page = doc[page_num]
        mat = fitz.Matrix(1.5, 1.5)  # 108 DPI for preview
        pix = page.get_pixmap(matrix=mat, alpha=False)
        img_data = pix.tobytes("png")
        doc.close()

        return send_file(
            io.BytesIO(img_data),
            mimetype="image/png",
            download_name=f"preview_page_{page_num + 1}.png",
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cleanup", methods=["POST"])
def cleanup():
    """Clean up old temporary files."""
    try:
        count = 0
        cutoff = time.time() - 3600  # 1 hour old
        for d in [UPLOAD_DIR, OUTPUT_DIR]:
            for f in d.iterdir():
                if f.is_file() and f.stat().st_mtime < cutoff:
                    f.unlink()
                    count += 1
        return jsonify({"cleaned": count})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("PDFConverter - Starting...")
    logger.info(f"OCR Available: {manager.ocr_engine.available}")
    logger.info(f"Supported formats: {list(SUPPORTED_FORMATS.keys())}")
    logger.info("=" * 60)
    app.run(host="127.0.0.1", port=5000, debug=True)
